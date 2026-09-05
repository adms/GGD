#!/usr/bin/env python3
"""傳說武器道具 → Excel。

來源全部是 repo 裡出貨的那一份（content/），不是任何手寫清單：
  · 池子    content/loot-tables/{legendary,ex-release,ex-origin}-weapons.json（三階）
  · 文件    content/items/<id>.json
  · 標記表  content/config/item-card.json

重跑：
    python3 -m venv "${TMPDIR:-/tmp}/xlsxenv" && "${TMPDIR:-/tmp}/xlsxenv/bin/pip" install openpyxl
    "${TMPDIR:-/tmp}/xlsxenv/bin/python" tools/economy/gen_legendary_xlsx.py

⚠️ 產出的 .xlsx **故意不進版控** —— owner 會在「修改需求」欄直接填字，
   追蹤它等於每次重跑都覆蓋掉他寫的東西。要最新的就重跑這支。
"""
import json, os, re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
OUT = os.path.join(ROOT, "docs", "傳說武器道具.xlsx")

CATEGORIES = {"武器", "法器", "飾品", "神器", "傳說", "任務", "夢幻", "積分獎勵", "防具", "鞋子"}

# ⚠️ 描述裡的方括號不全是「標記」。虛哭神去寫了
# 「[自身已損失的生命百分比數值(0~100)]」，那是一句說明不是標記。
# 判準：標記是短的中文詞（≤6 字）而且不含數字或括號。
def is_marker(t):
    return (len(t) <= 6 and not re.search(r"[0-9()（）%~／/、，。]", t)
            and re.fullmatch(r"[\u4e00-\u9fffA-Za-z\-]+", t) is not None)

STAT_ZH = {
    "maxHealth": "生命", "healthRegen": "每秒回血", "maxMana": "魔力", "manaRegen": "每秒回魔",
    "ad": "攻擊力", "ap": "AP", "armor": "防禦", "mr": "魔抗", "as": "攻速", "ms": "移速",
    "critChance": "暴擊率", "critDamage": "暴擊傷害", "cdr": "冷卻縮減", "lifesteal": "吸血",
    "spellVamp": "技能吸血", "range": "攻擊距離", "evasion": "迴避",
}
OP_ZH = {
    "flat": "+{v}", "pctAdd": "+{p}%（加進總和桶）", "pctMult": "×{m}（獨立相乘）",
    "percentOf": "= {p}% 的 {from}", "capRaise": "上限提升到 {v}",
}


def load(path):
    with open(os.path.join(ROOT, path), encoding="utf-8") as f:
        return json.load(f)


def split_description(desc):
    """→ (分類, 效能行[], 解說)"""
    lines = [l.rstrip() for l in (desc or "").split("\n")]
    category = lines[0].strip() if lines and lines[0].strip() in CATEGORIES else ""
    body = lines[1:] if category else lines[:]

    lore = ""
    idx = next((i for i, l in enumerate(body) if l.strip() == "解說"), None)
    if idx is not None:
        lore = "\n".join(x for x in body[idx + 1:] if x.strip())
        body = body[:idx]
    else:
        blanks = [i for i, l in enumerate(body) if not l.strip()]
        if blanks:
            cut = blanks[-1]
            tail = "\n".join(x for x in body[cut + 1:] if x.strip())
            # 只有當尾段看起來是敘述（不含 [標記]、不像數值行）才當成解說
            if tail and not tail.startswith("[") and not re.search(r"[+\-]\s*\d", tail.split("\n")[0]):
                lore, body = tail, body[:cut]
    eff = [l.strip() for l in body if l.strip() and l.strip() != "效能"]
    return category, eff, lore


def fmt_mod(m):
    stat = STAT_ZH.get(m.get("stat"), m.get("stat"))
    op = m.get("op")
    v = m.get("value")
    if op == "flat":
        # 比率型屬性用 % 讀比較直覺
        if m.get("stat") in ("lifesteal", "spellVamp", "cdr", "critChance", "evasion"):
            s = f"{stat} {round(v * 100, 4):+g}%"
        else:
            s = f"{stat} {v:+g}"
    elif op == "pctAdd":
        s = f"{stat} {round(v * 100, 4):+g}%（總和桶）"
    elif op == "pctMult":
        s = f"{stat} ×{round(1 + v, 4)}（獨立相乘）"
    elif op == "percentOf":
        src = STAT_ZH.get(m.get("from"), m.get("from")) or (
            "目前魔力" if m.get("fromResource") == "mp" else "目前生命")
        s = f"{stat} = {round(v * 100, 4)}% 的 {src}"
    elif op == "capRaise":
        s = f"{stat} 上限提升到 {v}"
    else:
        s = f"{stat} {op} {v}"
    req = m.get("requires") or {}
    if req.get("attackType"):
        s += f"（限{'近戰' if req['attackType'] == 'melee' else '遠戰'}）"
    if m.get("scopeSlot"):
        s += f"（只作用於 {m['scopeSlot']}）"
    return s


def fmt_hook(p):
    on = p.get("on", "?")
    zh = {
        "onBasicAttack": "普攻命中時", "onAbilityCast": "施放技能時", "onDamageTaken": "受到傷害時",
        "onKill": "擊殺時", "onInterval": "每隔一段時間", "onLevelUp": "升級時", "onHeal": "治療時",
    }.get(on, on)
    bits = [zh]
    if p.get("chance") is not None:
        bits.append(f"{round(p['chance'] * 100, 4)}% 機率")
    if p.get("internalCooldown") is not None:
        bits.append(f"內部冷卻 {p['internalCooldown']}s")
    if p.get("damageSource"):
        bits.append(f"僅限{'普攻' if p['damageSource'] == 'basic' else p['damageSource']}")
    if p.get("condition"):
        c = p["condition"]
        bits.append(f"條件：{c.get('kind')}"
                    + (f" {c.get('subject','')}.{c.get('stat','')} {c.get('op','')} {c.get('value','')}"
                       if c.get("kind") == "stat" else ""))
    kinds = "＋".join(e.get("kind", "?") for e in (p.get("effects") or []))
    return f"【{' · '.join(bits)}】→ {kinds}"


def fmt_aura(a):
    mods = "、".join(fmt_mod(m) for m in (a.get("modifiers") or []))
    who = {"enemy": "敵方", "ally": "友方", "all": "全體"}.get(a.get("affects"), a.get("affects"))
    return f"半徑 {a.get('radius')} 內的{who}：{mods or '（掛 hook）'}"


def special_fields(d):
    out = []
    if d.get("attributes"):
        a = d["attributes"]
        out.append("三圍：" + "、".join(
            f"{{'str':'力量','agi':'敏捷','int':'智力'}}[k] +{v}".format() if False else
            f"{ {'str':'力量','agi':'敏捷','int':'智力'}.get(k,k) } +{v}" for k, v in a.items()))
    if d.get("vision"):
        v = d["vision"]
        bits = []
        if v.get("trueSightRadius"):
            bits.append(f"看穿隱形，半徑 {v['trueSightRadius']}")
        if v.get("stealthFadeDelaySec") is not None:
            bits.append(f"隱身：無動作 {v['stealthFadeDelaySec']} 秒後再隱身")
        out.append("視野／隱身：" + "；".join(bits))
    if d.get("flight"):
        f = d["flight"]
        bits = [k for k, on in f.items() if on is True]
        zh = {"ignoreUnits": "無視單位碰撞", "ignoreObstacles": "無視障礙",
              "stayInsideBoundary": "仍受場地邊界限制"}
        out.append("飛行：" + "、".join(zh.get(b, b) for b in bits))
    if d.get("damageTypeOverride"):
        o = d["damageTypeOverride"]
        scope = {"basic": "普攻", "ability": "技能", "all": "全部"}.get(o.get("scope"), o.get("scope"))
        out.append(f"傷害轉換：{scope}傷害改判為「{o.get('becomes')}」"
                   + ("（特效仍用原本的類型）" if o.get("impactType") == "original" else ""))
    if d.get("block"):
        b = d["block"]
        types = "／".join({"physical": "物理", "magic": "魔法", "true": "真傷"}.get(t, t)
                          for t in (b.get("damageTypes") or []))
        s = f"格擋：{round(b.get('chance', 0) * 100, 4)}% 機率擋掉 {round(b.get('fraction', 1) * 100)}% 的{types}傷害"
        if b.get("lethalOnly"):
            s += "（只擋致命的那一擊）"
        if b.get("internalCooldown"):
            s += f"（內部冷卻 {b['internalCooldown']}s）"
        out.append(s)
    if d.get("critStrike"):
        c = d["critStrike"]
        s = f"暴擊：{round(c.get('chance', 0) * 100, 4)}% 機率造成 {c.get('damageMult')} 倍傷害"
        if c.get("lifestealFraction") is not None:
            mode = "取代原本吸血" if c.get("lifestealMode") == "replace" else "與原本吸血相加"
            s += f"；暴擊時吸血 {round(c['lifestealFraction'] * 100)}%（{mode}）"
        if c.get("empowers"):
            s += f"；作用範圍 {c['empowers']}"
        out.append(s)
    if d.get("requiresAttackType"):
        out.append(f"限定：{'近戰' if d['requiresAttackType'] == 'melee' else '遠戰'}英雄才能裝備")
    return "\n".join(out)


def fmt_set(s):
    need = s.get("requiredPieces") or len(s.get("pieces", []))
    mods = "、".join(fmt_mod(m) for m in (s.get("modifiers") or []))
    return f"{s.get('name') or s.get('id')}：集滿 {need}/{len(s.get('pieces', []))} 件 → {mods}\n成員：{'、'.join(s.get('pieces', []))}"


# ---------------------------------------------------------------- 讀資料
# ⭐ owner 2026-08-18 把上架寶具切成三階（EX / [EX解放] / [EX∅ 根源]）。
# ⛔ 件數不再寫死 —— 讀三張檔，一件寶具只屬於一個池。
POOL_TABLES = ["legendary-weapons", "ex-release-weapons", "ex-origin-weapons"]
ids = []
for _t in POOL_TABLES:
    _lt = load(f"content/loot-tables/{_t}.json")
    ids += [e.get("itemId") or e.get("id") if isinstance(e, dict) else e
            for e in (_lt.get("entries") or _lt.get("items") or [])]
docs = {}
for fn in os.listdir(os.path.join(ROOT, "content/items")):
    if fn.endswith(".json") and not fn.startswith("_"):
        d = load("content/items/" + fn)
        docs[d["id"]] = d
card = load("content/config/item-card.json")
markers = card.get("markers", {})
cat_zh = {"stat": "常駐屬性", "active": "會觸發的事件", "passive": "常駐特性", "debuff": "對敵減益"}

# ---------------------------------------------------------------- 活頁簿
wb = Workbook()

HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ZEBRA = PatternFill("solid", fgColor="F2F5FA")


def style_sheet(ws, widths, wrap_cols, freeze="A2", row_height=None):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for c in ws[1]:
        c.fill, c.font = HEAD_FILL, HEAD_FONT
        c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    for r in range(2, ws.max_row + 1):
        if row_height:
            ws.row_dimensions[r].height = row_height
        for ci in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=ci)
            cell.border = BORDER
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=(get_column_letter(ci) in wrap_cols),
            )
            if r % 2 == 0:
                cell.fill = ZEBRA


# ============================ Sheet 1：傳說道具 ============================
ws = wb.active
ws.title = "傳說道具"
HEADERS = [
    "#", "名稱", "修改需求（你填）", "ID", "階級", "分類",
    "能力說明（效能）", "標記", "常駐屬性加成", "AP 加成", "觸發（[xxx時]）",
    "光環", "套裝", "特殊機制欄位", "tags", "合成", "解說（風味）", "完整 JSON",
]
ws.append(HEADERS)

for n, iid in enumerate(ids, start=1):
    d = docs.get(iid)
    if not d:
        continue
    category, eff, lore = split_description(d.get("description", ""))
    tags_in_text = []
    for m in re.finditer(r"\[([^\]]+)\]", d.get("description", "")):
        t = m.group(1).strip()
        if is_marker(t) and t not in tags_in_text:
            tags_in_text.append(t)
    marker_col = "\n".join(
        f"[{t}] — {cat_zh.get(markers.get(t), '（未登記分類）')}" for t in tags_in_text)
    mods = [fmt_mod(m) for m in (d.get("modifiers") or [])]
    ap_bits = []
    for m in (d.get("modifiers") or []):
        if m.get("stat") == "ap":
            ap_bits.append(fmt_mod(m))
    for s in (d.get("sets") or []):
        for m in (s.get("modifiers") or []):
            if m.get("stat") == "ap":
                ap_bits.append(f"（套裝）{fmt_mod(m)}")
    recipe = d.get("recipe")
    ws.append([
        n,
        d.get("name", ""),
        "",
        d.get("id", ""),
        d.get("tier", ""),
        category,
        "\n".join(eff),
        marker_col,
        "\n".join(mods),
        "\n".join(ap_bits),
        "\n".join(fmt_hook(p) for p in (d.get("passive") or [])),
        "\n".join(fmt_aura(a) for a in (d.get("auras") or [])),
        "\n".join(fmt_set(s) for s in (d.get("sets") or [])),
        special_fields(d),
        "、".join(d.get("tags") or []),
        (("" if d.get("craftRole") in (None, "none") else {"final": "最終成品", "quest": "任務獎勵"}.get(d["craftRole"], d["craftRole"]))
         + (f"｜材料：{json.dumps(recipe, ensure_ascii=False)}" if recipe else "")),
        lore,
        json.dumps(d, ensure_ascii=False, indent=2),
    ])

style_sheet(
    ws,
    widths=[4, 16, 18, 14, 6, 10, 52, 30, 34, 28, 46, 34, 40, 44, 14, 22, 50, 70],
    wrap_cols=set("BCFGHIJKLMNQR"),
)
for r in range(2, ws.max_row + 1):
    ws.cell(row=r, column=2).font = Font(bold=True)
    ws.cell(row=r, column=18).font = Font(name="Menlo", size=9)
    ws.cell(row=r, column=18).alignment = Alignment(vertical="top", wrap_text=False)

# ============================ Sheet 2：標記對照表 ============================
ws2 = wb.create_sheet("標記對照表")
ws2.append(["標記", "分類", "分類意義", "用到的道具數", "哪些道具"])
usage = {}
for iid in ids:
    d = docs.get(iid)
    if not d:
        continue
    for m in re.finditer(r"\[([^\]]+)\]", d.get("description", "")):
        t = m.group(1).strip()
        if is_marker(t):
            usage.setdefault(t, set()).add(d["name"])
for t in sorted(usage, key=lambda k: (-len(usage[k]), k)):
    ws2.append([
        f"[{t}]",
        markers.get(t, "（未登記）"),
        cat_zh.get(markers.get(t), "⚠️ 卡片會落到 unknownCategory"),
        len(usage[t]),
        "、".join(sorted(usage[t])),
    ])
style_sheet(ws2, widths=[14, 12, 30, 12, 70], wrap_cols=set("CE"))

# ============================ Sheet 3：機制欄位覆蓋 ============================
ws3 = wb.create_sheet("機制欄位覆蓋")
SURF = [
    ("modifiers", "常駐屬性"), ("passive", "觸發事件"), ("auras", "光環"), ("sets", "套裝"),
    ("attributes", "三圍"), ("vision", "視野/隱身"), ("flight", "飛行"),
    ("damageTypeOverride", "傷害轉換"), ("block", "格擋"), ("critStrike", "暴擊"),
    ("recipe", "合成"), ("requiresAttackType", "攻擊型態限定"),
]
ws3.append(["名稱", "ID"] + [f"{zh}\n({k})" for k, zh in SURF] + ["實作面數"])
for iid in ids:
    d = docs.get(iid)
    if not d:
        continue
    row = [d["name"], d["id"]]
    cnt = 0
    for k, _ in SURF:
        v = d.get(k)
        on = bool(v) and (len(v) > 0 if isinstance(v, (list, dict)) else True)
        row.append("●" if on else "")
        cnt += 1 if on else 0
    row.append(cnt)
    ws3.append(row)
style_sheet(ws3, widths=[16, 14] + [11] * len(SURF) + [10], wrap_cols=set())
for r in range(2, ws3.max_row + 1):
    for c in range(3, 3 + len(SURF)):
        ws3.cell(row=r, column=c).alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 46

# ============================ Sheet 4：說明 ============================
ws4 = wb.create_sheet("這份表怎麼讀", 0)
NOTES = [
    ["這份表是什麼", ""],
    ["", "content/ 出貨檔直接產生的，不是手寫清單。改了內容重跑一次就是最新的。"],
    ["", ""],
    ["來源（全部可追）", ""],
    [f"池子（哪 {len(ids)} 件會出現在三選一）", "content/loot-tables/{" + ",".join(POOL_TABLES) + "}.json"],
    ["每一件的資料", "content/items/<id>.json"],
    ["標記→分類", "content/config/item-card.json 的 markers"],
    ["", ""],
    ["分頁", ""],
    ["傳說道具", "主表。一列一件，含能力說明、標記、屬性、觸發、套裝、完整 JSON。"],
    ["標記對照表", "所有 [標記] 用了幾次、分到哪一類。「（未登記）」＝卡片畫不出正確顏色。"],
    ["機制欄位覆蓋", "每一件用到哪些實作欄位。空白列＝這件道具沒有任何實作（目前 0 件）。"],
    ["", ""],
    ["欄位要注意的", ""],
    ["修改需求（你填）", "留空給你寫。跟 items.csv 一樣的用法：填中文需求，我來翻成 JSON。"],
    ["常駐屬性加成 ── 相加 vs 相乘", "「總和桶」= pctAdd，同類全部先相加再乘一次（+100% 與 +300% ⇒ ×5.0，不是 ×8.0）。"],
    ["", "「獨立相乘」= pctMult，每一條各乘一次。兩者是不同的桶。"],
    ["AP 加成", "從 modifiers 與套裝獎勵抽出來的快查欄，方便盤法系裝。"],
    ["觸發（[xxx時]）", "格式是【什麼時候 · 機率 · 內部冷卻 · 條件】→ 會發生什麼效果。"],
    ["特殊機制欄位", "三圍/視野/飛行/傷害轉換/格擋/暴擊 —— 這六個是頂層欄位，不在 modifiers 裡。"],
    ["", "⚠️ 只看 modifiers 會漏掉它們（2026-08-10 我就是這樣看漏的）。"],
    ["完整 JSON", "原始文件全文，可以直接貼回 content/items/<id>.json。"],
    ["", ""],
    ["JSON 可組性", f"每一件「需不需要額外程式」。今天 {len(ids)} 件全部都是「不用」——"],
    ["", "機制全部由現有的 item@1 欄位組出來。這一欄有一天變成「需要」時，說明欄會寫出卡在哪一個機制。"],
    ["機制詞彙表", "引擎目前提供的 12 個機制欄位各自能表達什麼、誰在用。"],
    ["", "這一頁就是「JSON 能組出什麼」的完整答案 —— 要新機制才寫得出來的東西，不會出現在這裡。"],
]
ws4.append(["項目", "說明"])
for a, b in NOTES:
    ws4.append([a, b])
style_sheet(ws4, widths=[34, 96], wrap_cols=set("B"))
for r in range(2, ws4.max_row + 1):
    if ws4.cell(row=r, column=2).value == "":
        ws4.cell(row=r, column=1).font = Font(bold=True, color="1F3864", size=12)


# ============================ Sheet 5：JSON 可組性 ============================
ws5 = wb.create_sheet("JSON 可組性")
ws5.append(["#", "名稱", "ID", "效能行數", "用到的實作面", "面數",
            "需要額外程式嗎", "說明"])
NEED_CODE = {}  # id -> 說明；今天是空的，欄位留著是為了「有一天不是空的」時看得見
for n, iid in enumerate(ids, start=1):
    d = docs.get(iid)
    if not d:
        continue
    used = [zh for k, zh in SURF if d.get(k) and (len(d[k]) > 0 if isinstance(d[k], (list, dict)) else True)]
    _, eff, _ = split_description(d.get("description", ""))
    why = NEED_CODE.get(iid)
    ws5.append([
        n, d["name"], d["id"], len(eff), "、".join(used), len(used),
        "❌ 需要" if why else "✅ 不用",
        why or "全部機制都由現有的 item@1 欄位組出來（modifiers／passive／auras／sets ＋ 六個頂層機制欄位）。",
    ])
style_sheet(ws5, widths=[4, 16, 14, 10, 40, 6, 12, 62], wrap_cols=set("EH"))

# ============================ Sheet 6：機制詞彙表 ============================
ws6 = wb.create_sheet("機制詞彙表")
ws6.append(["機制欄位", "住在哪", "它能表達什麼", f"這 {len(ids)} 件裡誰在用"])
VOCAB = [
    ("modifiers", "item@1.modifiers",
     "常駐屬性。op 決定怎麼疊：flat 相加／pctAdd 進**同一個總和桶**（+100% 與 +300% ⇒ ×5.0，不是 ×8.0）／"
     "pctMult 各自相乘／percentOf 取另一條屬性或當下資源的百分比／capRaise 抬高上限。"),
    ("passive", "item@1.passive",
     "[xxx時] 觸發。on = 普攻命中／施放技能／受到傷害／擊殺／每隔一段時間；再配 chance 機率、"
     "internalCooldown 內部冷卻、condition 條件葉。effects 是效果陣列，順序就是執行順序。"),
    ("auras", "item@1.auras",
     "光環。半徑 × 影響誰（敵／友／全體）× 一組 modifiers 或 hook。"),
    ("sets", "item@1.sets",
     "套裝。pieces 列出成員、requiredPieces 決定湊幾件生效、modifiers 是獎勵（只發一次，不是每件一份）。"),
    ("attributes", "item@1.attributes", "力/敏/智三圍。走 sim/stats/attrSources.ts，會再換算成 AD／攻速／AP。"),
    ("vision", "item@1.vision", "看穿隱形（trueSightRadius）／自身隱身的再隱身延遲（stealthFadeDelaySec）。"),
    ("flight", "item@1.flight", "飛行形態：無視單位碰撞／無視障礙／是否仍受場地邊界限制。"),
    ("damageTypeOverride", "item@1.damageTypeOverride",
     "傷害轉換。scope 選普攻／技能／全部，becomes 改判成 true（真傷）等，impactType 可保留原本特效類型。"),
    ("block", "item@1.block", "格擋。擋哪幾種傷害 × 機率 × 擋掉幾成 × 是否只擋致命那一擊 × 內部冷卻。"),
    ("critStrike", "item@1.critStrike",
     "暴擊。機率 × 倍率 × 暴擊時吸血比例（取代或相加）× 作用範圍。"),
    ("requiresAttackType", "item@1.requiresAttackType", "限定近戰或遠戰英雄才能裝備。"),
    ("recipe", "item@1.recipe", "合成材料。"),
]
for key, home, what in VOCAB:
    users = [docs[i]["name"] for i in ids
             if docs.get(i) and docs[i].get(key)
             and (len(docs[i][key]) > 0 if isinstance(docs[i][key], (list, dict)) else True)]
    ws6.append([key, home, what, f"{len(users)} 件：" + "、".join(users[:8]) + ("…" if len(users) > 8 else "")])
style_sheet(ws6, widths=[22, 30, 76, 56], wrap_cols=set("CD"))

wb.save(OUT)
print("寫出：", OUT)
print("道具列數：", ws.max_row - 1)
print("標記數：", ws2.max_row - 1)
