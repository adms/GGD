#!/usr/bin/env python3
"""
⭐【把「還沒重製的技能」倒成一份可以直接填的 Excel】—— owner 2026-08-13：

    「最新的英雄列表 md 請你輸出成 excel 讓我可以修改剩下的英雄技能
      繼續技能重製計畫吧」

用法：
    python3 tools/skill-remake/export_xlsx.py            # 寫到 docs/skill-remake/
    python3 tools/skill-remake/export_xlsx.py --out X    # 指定輸出檔

⚠️ 這支是**產生器**不是一次性腳本：每做完一批就重跑一次，已重製的會自動離開
   「待重製」分頁。⛔ 不要手改產出的 xlsx 再拿它當來源 —— 來源永遠是
   `content/abilities/*.json` 與 `tools/skill-remake/batch1.py`。

判定「已重製」的方式與 `stamp_provenance.py` **同一條**：讀 `batch1.py` 的 `HERO`
對照表，⛔ 不是手抄一份 id 清單（那會過期）。

三個分頁：
  · 待重製      —— 一列一支技能，兩欄留白給 owner 填（新版技能名 / 新版技能說明）
  · 已完成90支  —— 對照用，讓 owner 看得到自己上一批寫的**格式**
  · 填寫說明    —— 規則 + 標籤詞彙（詞彙是從那 90 支**量出來**的，⛔ 不是我發明的）
"""
from __future__ import annotations

import argparse
import collections
import glob
import json
import os
import re
import sys

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
GEN = os.path.join(ROOT, "tools", "skill-remake", "batch1.py")
SLOT_ORDER = {"passive": 0, "q": 1, "w": 2, "e": 3, "r": 4, "ex": 5}
SLOT_LABEL = {"passive": "天生技", "q": "Q", "w": "W", "e": "E", "r": "R", "ex": "EX"}


def remade_prefixes() -> set[str]:
    src = open(GEN, encoding="utf-8").read()
    m = re.search(r"^HERO\s*=\s*\{(.*?)^\}", src, re.S | re.M)
    if m is None:
        raise SystemExit("⛔ 在 batch1.py 找不到 HERO 對照表 —— 它改形狀了，先看過再跑")
    return set(re.findall(r'"(godie-[a-z0-9]+)"', m.group(1)))


def live_champion_ids() -> set[str]:
    """營運白名單（gitignore 的機器狀態，缺席時回空集合 —— 那只是少一欄提示）。"""
    p = os.path.join(ROOT, "data", "curation", "whitelist.json")
    try:
        return set(json.load(open(p, encoding="utf-8")).get("champions") or [])
    except Exception:
        return set()


def load(collection: str) -> dict[str, dict]:
    out = {}
    for f in glob.glob(os.path.join(ROOT, "content", collection, "*.json")):
        if os.path.basename(f).startswith("_"):
            continue
        d = json.load(open(f, encoding="utf-8"))
        out[d["id"]] = d
    return out


def hero_no(name: str) -> str:
    """「70-01 伸卡球」→「70」。抓不到就空字串（排序時排最後）。"""
    m = re.match(r"\s*(\d{2})-", name or "")
    return m.group(1) if m else ""


def ability_no(name: str) -> str:
    m = re.match(r"\s*(\d{2}-\d{2,3})", name or "")
    return m.group(1) if m else ""


def first_line_tags(desc: str) -> str:
    return " ".join(f"[{t}]" for t in re.findall(r"\[([^\]]+)\]", (desc or "").split("\n")[0]))


def effect_summary(doc: dict) -> str:
    """現有效果的**形狀**摘要（kind 次數）。⛔ 不是數值 —— 那是 owner 要重寫的東西。"""
    kinds: collections.Counter = collections.Counter()

    def walk(node):
        if isinstance(node, list):
            for x in node:
                walk(x)
        elif isinstance(node, dict):
            k = node.get("kind")
            if isinstance(k, str):
                kinds[k] += 1
            for v in node.values():
                walk(v)

    walk(doc.get("effects"))
    walk(doc.get("passive"))
    return " · ".join(f"{k}×{n}" if n > 1 else k for k, n in kinds.most_common()) or "（空）"


def num_list(v) -> str:
    if v is None:
        return ""
    if isinstance(v, list):
        return "/".join(str(x) for x in v)
    return str(v)


def rows_for(prefixes: set[str], want_remade: bool, abilities, champs, live):
    out = []
    for aid, doc in abilities.items():
        head = aid.split(".")[0]
        if (head in prefixes) != want_remade:
            continue
        champ = champs.get(head) or {}
        slot = aid.split(".")[-1]
        name = doc.get("name") or ""
        out.append(
            {
                "英雄編號": hero_no(name) or hero_no((champ.get("name") or "")),
                "英雄名": champ.get("name") or head,
                "英雄ID": head,
                "營運中": "✔" if head in live else "",
                "槽位": SLOT_LABEL.get(slot, slot),
                "技能編號": ability_no(name),
                "技能ID": aid,
                "現有技能名": name,
                "新版技能名（填這裡）": "",
                "新版技能說明（填這裡）": "",
                "現有說明（參考）": doc.get("description") or "",
                "現有標籤": first_line_tags(doc.get("description") or ""),
                "冷卻": num_list(doc.get("cooldown")),
                "魔耗": num_list(doc.get("manaCost")),
                "射程": num_list(doc.get("range")),
                "現有效果形狀": effect_summary(doc),
                "_sort": (
                    hero_no(name) or "99",
                    head,
                    SLOT_ORDER.get(slot, 9),
                ),
            }
        )
    out.sort(key=lambda r: r["_sort"])
    for r in out:
        del r["_sort"]
    return out


def tag_vocabulary(abilities, prefixes) -> list[tuple[str, int]]:
    """owner 在那 90 支第一行**真的用過**的標籤。⛔ 不是我編一份給他。"""
    c: collections.Counter = collections.Counter()
    for aid, doc in abilities.items():
        if aid.split(".")[0] not in prefixes:
            continue
        for t in re.findall(r"\[([^\]]+)\]", (doc.get("description") or "").split("\n")[0]):
            c[t] += 1
    return c.most_common()


HOWTO = [
    ("這份檔案是什麼", "「剩下的英雄技能」的填寫表。每一列是一支技能，你只要填**兩欄**："
     "『新版技能名』與『新版技能說明』。其餘欄位都是參考用的現況，⛔ 不用動。"),
    ("留白就是「這一支不改」", "沒填的列會被跳過，照現在的樣子繼續出貨。想分批做就分批填。"),
    ("", ""),
    ("⭐ 說明要怎麼寫", "照你上一批 90 支的格式（見『已完成90支』分頁）：第一行是方括號標籤，"
     "接著冷卻／消耗／距離／半徑，空一行，然後是內文。"),
    ("⭐「」裡面是角色對白", "**不是效果**。你寫在「」裡的台詞我不會當成機制去實作。"
     "（44-04 心臟麻痺的「在35秒後宣布勝利吧」曾經被誤讀成一支有 35 秒延遲的技能。）"),
    ("⭐ 內文贏過標籤", "標籤與內文打架時**以內文為準**，而且我會把標籤改成符合內文，"
     "⛔ 不會留兩個互相打架的來源。"),
    ("", ""),
    ("數值怎麼寫", "逐階用斜線：`造成250/350/450/550傷害`、`消耗MP150/190/230/240`。"
     "只有一階就寫一個數字。"),
    ("範圍不要寫數字", "寫『[範圍]』『[周圍]』就好 —— 半徑走四級距（小/中/大/超大），"
     "引擎會翻譯。你堅持要指定半徑也可以，寫『有效半徑6.05』。"),
    ("吟唱時間", "0.06~4.00 秒之間（你 2026-08-13 定的）。不寫的話我用公式推。"),
    ("", ""),
    ("填完之後", "把檔案給我，我把它併進 `tools/skill-remake/batch1.py`（那是 90 支的**唯一來源**），"
     "重跑產生器 → JSON、文件、對外契約一次全部更新。"),
    ("⛔ 不要直接改 JSON", "`content/abilities/*.json` 是產生器的輸出，手改會在下一次重跑時被蓋掉。"),
]


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=os.path.join(ROOT, "docs", "skill-remake", "GGD-技能重製-待填.xlsx"))
    args = ap.parse_args()

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("⛔ 需要 openpyxl：python3 -m pip install --user openpyxl", file=sys.stderr)
        return 2

    prefixes = remade_prefixes()
    abilities = load("abilities")
    champs = load("champions")
    live = live_champion_ids()

    todo = rows_for(prefixes, False, abilities, champs, live)
    done = rows_for(prefixes, True, abilities, champs, live)

    wb = Workbook()
    HEAD_FILL = PatternFill("solid", fgColor="1F3864")
    FILL_ME = PatternFill("solid", fgColor="FFF2CC")
    HEAD_FONT = Font(color="FFFFFF", bold=True)

    def sheet(title, rows, fill_cols):
        ws = wb.create_sheet(title)
        if not rows:
            ws["A1"] = "（沒有資料）"
            return ws
        cols = list(rows[0].keys())
        ws.append(cols)
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = HEAD_FILL
            cell.font = HEAD_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for r in rows:
            ws.append([r[c] for c in cols])
        widths = {
            "英雄編號": 9, "英雄名": 22, "英雄ID": 14, "營運中": 8, "槽位": 8,
            "技能編號": 10, "技能ID": 20, "現有技能名": 22,
            "新版技能名（填這裡）": 22, "新版技能說明（填這裡）": 60,
            "現有說明（參考）": 60, "現有標籤": 24, "冷卻": 14, "魔耗": 16,
            "射程": 10, "現有效果形狀": 34,
        }
        for i, c in enumerate(cols, start=1):
            ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 16)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(cols)):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for name in fill_cols:
            if name in cols:
                i = cols.index(name) + 1
                for r in range(2, ws.max_row + 1):
                    ws.cell(row=r, column=i).fill = FILL_ME
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        return ws

    sheet("待重製", todo, ["新版技能名（填這裡）", "新版技能說明（填這裡）"])
    # 已完成那一批：把 owner 自己的規格放進「說明」欄，⛔ 那一批不需要留白
    for r in done:
        r["新版技能名（填這裡）"] = r["現有技能名"]
        r["新版技能說明（填這裡）"] = r["現有說明（參考）"]
    sheet("已完成90支", done, [])

    ws = wb.create_sheet("填寫說明")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 96
    ws["A1"], ws["B1"] = "項目", "說明"
    for c in ("A1", "B1"):
        ws[c].fill = HEAD_FILL
        ws[c].font = HEAD_FONT
    for k, v in HOWTO:
        ws.append([k, v])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.append([])
    ws.append(["標籤詞彙（你在 90 支裡真的用過的）", ""])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    vocab = tag_vocabulary(abilities, prefixes)
    line, buf = [], []
    for t, n in vocab:
        buf.append(f"[{t}]×{n}")
        if len(buf) == 8:
            line.append(" ".join(buf))
            buf = []
    if buf:
        line.append(" ".join(buf))
    for s in line:
        ws.append(["", s])
        ws.cell(row=ws.max_row, column=2).alignment = Alignment(wrap_text=True)

    del wb["Sheet"]
    os.makedirs(os.path.dirname(args.out), exist_ok=True)
    wb.save(args.out)
    heroes = len({r["英雄ID"] for r in todo})
    print(f"寫出 {args.out}")
    print(f"  待重製：{len(todo)} 支 / {heroes} 位英雄（其中營運中 "
          f"{len({r['英雄ID'] for r in todo if r['營運中']})} 位）")
    print(f"  已完成：{len(done)} 支 / {len({r['英雄ID'] for r in done})} 位英雄")
    print(f"  標籤詞彙：{len(vocab)} 種")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
